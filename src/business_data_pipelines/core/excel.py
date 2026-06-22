from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlrd


def read_two_row_header_sheet(path: Path, *, data_start_row: int = 4) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook.active
    row1 = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    row2 = [sheet.cell(2, col).value for col in range(1, sheet.max_column + 1)]
    headers = [row2[i] or row1[i] for i in range(sheet.max_column)]

    records: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if all(value is None for value in row):
            continue
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return records


def read_single_row_header_sheet(
    path: Path,
    *,
    header_row: int = 1,
    data_start_row: int = 2,
) -> tuple[list[str], list[dict[str, Any]]]:
    if _is_xls(path):
        return _read_single_row_header_xls(
            path,
            header_row=header_row,
            data_start_row=data_start_row,
        )
    return _read_single_row_header_xlsx(
        path,
        header_row=header_row,
        data_start_row=data_start_row,
    )


def _is_xls(path: Path) -> bool:
    with path.open("rb") as file:
        return file.read(8).startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")


def _read_single_row_header_xlsx(
    path: Path,
    *,
    header_row: int,
    data_start_row: int,
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook.active
    headers = [
        _normalize_header(sheet.cell(header_row, col).value)
        for col in range(1, sheet.max_column + 1)
    ]
    records: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if all(value is None for value in row):
            continue
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return headers, records


def _read_single_row_header_xls(
    path: Path,
    *,
    header_row: int,
    data_start_row: int,
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = xlrd.open_workbook(str(path))
    sheet = workbook.sheet_by_index(0)
    headers = [_normalize_header(sheet.cell_value(header_row - 1, col)) for col in range(sheet.ncols)]
    records: list[dict[str, Any]] = []
    for row_index in range(data_start_row - 1, sheet.nrows):
        values = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
        if all(value in {None, ""} for value in values):
            continue
        records.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    return headers, records


def _normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()
